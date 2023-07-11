import re
import warnings
import openpyxl
import pandas as pd
from logs import log

from user_settings.user_exceptions import WorkbookFilterError, NotFoundColumns


class BankFile1:
    BANK_NAMES = ['СБЕР', 'Альфа Банк', 'Совкомбанк', 'Дом РФ', 'Другой банк', 'Олд Банк', 'Яуза Банк',
                  'Остафьево Банк', 'Cпутник Банк']
    COLUMNS_FOR_BANKS = {
        'СБЕР': (['Депонент', 'Номер ДДУ', 'Исходящий остаток', 'Номер счёта эскроу'],[0, 1, 2], 0),
        'Альфа Банк': (['ФИО плательщика', 'Номер ДДУ', 'Остаток на счете эскроу', 'Номер счета эскроу'],[0], 0),
        'Другой банк': (['ФИО депонента', 'Данные дог.об участии сч.(ДДУ)', 'Текущий остаток на счете эскроу на дату отчета', 'Номер счета эскроу'],[0], 0),
        'Яуза Банк': (['Депонент', 'Номер ДДУ', 'Исходящий остаток на дату операции, RUB', 'Номер счёта эскроу'],[0], 0),
        'Совкомбанк': (['Наименование Депонента', 'Номер ДДУ', 'Исходящий остаток депонируемой суммы, RUB', 'Номер счета эскроу'],[0,1], 1),
        'Остафьево Банк': (['ФИО Плательщика/Наименование', 'Номер ДДУ', 'Сумма по ДДУ, руб.', 'Номер счета эскроу'],[0], 0),
        'Олд Банк': (['Депонент (ФИО)', 'Номер договора основания', 'Исходящий остаток', 'Счет-эскроу'],[0], 0),
        'Дом РФ': (['Депонент', '№ договора ДДУ', 'Остаток денежных средств на конец периода,  руб.', 'Номер счета'],[0], 0),
        'Cпутник Банк': ([8, 7, 14, 6], [0, 1, 2])
    }
    NEW_COLUMNS = ['Контрагент', 'Договор', 'Сальдо', 'Номер счета']

    REGEX_CONTRACT_PATTERN = {3: r'\w+ \w*\-\d\.?\d*\/\d*\.?\d?', 2: r'\w*\-\d\.?\d*\/\d*\.?\d?', 4: r'\w*\-?\/\d\.?\d*\/\d*',
                     5: r'[А-Яа-я]+\-\d*[.-/]?\d*', 6: r'[А-Яа-я]+\/[А-Яа-я]+\/\d+\-\d+', 7: r'[А-Яа-я]+\/\d+\-\d+', 8: r'\w*\/\d\.?\d?\-?\d*',
                              1: r'\w+[-/]\d+[-/]\d\.?\d?[-/]\d+[-/].*'}

    def __init__(self, user_data_dict):
        self.name_bank_file = user_data_dict['bank_file']
        self.name_bank_folder = user_data_dict['bank_folder']
        if len(user_data_dict['bank_name'])>0:
            self.bank_name = user_data_dict['bank_name'][0]
        else:
            self.bank_name = ''
        self.is_one_file = user_data_dict['single']
        self.file_to_bank = user_data_dict['file_to_bank']
        self.document_sum = 0
        self.count_documents = 0
        self.df = self.create_and_edit_bank_dataframe()


    def create_and_edit_bank_dataframe(self):
        result = pd.DataFrame()
        if self.name_bank_file != '':
            short_file_name = self.name_bank_file[self.name_bank_file.rfind('/') + 1:]
            result = self.edit(self.bank_name, self.name_bank_file, short_file_name)
            self.count_documents+=1
        else:
            temp_bank_data = pd.DataFrame(columns=self.NEW_COLUMNS)
            for key, value in self.file_to_bank.items():
                part_bank_data = self.edit(value[1], value[0], key)
                temp_bank_data = pd.concat([temp_bank_data, part_bank_data])
                self.count_documents += 1
            result = temp_bank_data
        result = self.set_index_on_df(result)
        return result

    def edit(self, bank_name, file_name, short_file_name):
        log.info(f'Считывание файла {short_file_name}')
        data = pd.DataFrame()
        try:
            with warnings.catch_warnings(record=True):
                warnings.simplefilter("always")
                wb = openpyxl.load_workbook(file_name)
            if bank_name == 'Дом РФ':
                raw_data = wb[wb.sheetnames[1]]
            else:
                raw_data = wb[wb.sheetnames[0]]
            data = self.edit_bank_data(raw_data, bank_type=bank_name)
        except WorkbookFilterError as exp:
            raise WorkbookFilterError(self.name_bank_file)
        else:
            wb.close()
            return data

    def edit_bank_data(self, raw_bank_data, bank_type):
        bank_data = self.remove_na_rows(raw_bank_data, bank_type)
        try:
            bank_data = self.select_bank_columns(bank_data, bank_type)
        except NotFoundColumns as exp:
            log.exception(NotFoundColumns)
            raise NotFoundColumns
        bank_data = self.rename_columns(bank_data)
        return bank_data

    def rename_columns(self, df):
        string_columns = ['Контрагент', 'Договор', 'Номер счета']
        for column in string_columns:
            df[column] = df[column].apply(lambda x: str(x).strip())
        df["Сальдо"] = pd.to_numeric(df["Сальдо"], errors='ignore')
        df = df.groupby(['Контрагент', 'Договор', 'Номер счета'], as_index=False).agg(sum)
        self.document_sum += self.sum_amount(df)
        df['Сальдо'] = df['Сальдо'].apply(str)
        df['Контрагент'] = df['Контрагент'].apply(lambda x: self.edit_bank_agent(x))
        return df

    def remove_na_rows(self, raw_bank_data, bank):
        row_list = list(raw_bank_data.values)
        bank_data = pd.DataFrame(row_list)
        col_indexes = []
        if bank == 'Совкомбанк':
            for i, col in enumerate(bank_data.columns):
                if sum(map(lambda x: x is None, bank_data[col])) > sum(map(lambda x: x is not None, bank_data[col])):
                    col_indexes.append(i)
            bank_data.drop(col_indexes, axis=1, inplace=True)
        row_indexes = []
        for i in range(len(bank_data)):
            if sum(map(lambda x: x is None, bank_data.iloc[i])) > sum(map(lambda x: x is not None, bank_data.iloc[i])):
                if bank == 'Яуза Банк' and bank_data.iloc[i][1] is None and bank_data.iloc[i][13] is not None:
                    continue
                row_indexes.append(i)
        bank_data = bank_data.drop(index=row_indexes)
        bank_data.reset_index(inplace=True)

        if bank == 'Cпутник Банк':
            bank_data.drop(index=self.COLUMNS_FOR_BANKS[bank][1], inplace=True)
        else:
            bank_data.columns = bank_data.iloc[self.COLUMNS_FOR_BANKS[bank][2]]
            try:
                bank_data.drop(index=self.COLUMNS_FOR_BANKS[bank][1], inplace=True)
            except KeyError as kerr:
                log.info('Данный файл пуст')
                bank_data.drop(index=[0], inplace=True)
            bank_data.drop(columns=bank_data.columns[0], axis=1, inplace=True)
        return bank_data


    def select_bank_columns(self, df, bank):
        if bank == 'Олд Банк':
            df.columns = df.columns.map(lambda x: x.replace('\n', ' ') if type(x) == str else x)
        df.sort_values(by=self.COLUMNS_FOR_BANKS[bank][0][1], inplace=True)
        df = df[self.COLUMNS_FOR_BANKS[bank][0]]
        df.columns = self.NEW_COLUMNS
        if bank == 'Альфа Банк':
            df['Сальдо'] = df['Сальдо'].apply(lambda x: float(str(x).replace(u'\xa0', u'').replace(',', '.')))
        elif bank=='Яуза Банк':
            df = self.fill_na_bank_data(df)
        elif bank == 'Дом РФ':
            df = df[df['Сальдо'].map(lambda x: type(x) != str)]
        df['Договор (полный)'] = df['Договор'].apply(lambda x: str(x).upper())
        df['Договор'] = df['Договор'].apply(lambda x: str(x).upper())
        return df

    def find_query_for_bank(self, contract):
        if contract.find('/') != -1 and contract.find('-') != -1:
            split_ctr = re.split(r'\-|\/', contract)
            candidat = split_ctr[1]
            if all(map(str.isalpha, candidat)) and len(split_ctr) >= 4:
                return split_ctr[2]
            else:
                return candidat
        elif contract.count('/') == 2:
            split_ctr = re.split(r'\/', contract)
            if len(split_ctr) >= 3:
                return split_ctr[1]
        else:
            return 'бн'

    def find_house_for_bank(self, contract):
        if contract.find('/') != -1 and contract.find('-') != -1:
            split_ctr = re.split(r'\-|\/', contract)
            candidat = split_ctr[2]
            prev_canidat = split_ctr[1]
            if all(map(str.isalpha, prev_canidat)) and len(split_ctr) >= 4:
                return split_ctr[3]
            else:
                return candidat
        elif contract.count('/') == 2:
            split_ctr = re.split(r'\/', contract)
            if len(split_ctr) >= 3:
                return split_ctr[2]
        else:
            return 'бн'



    def fill_na_bank_data(self, df):
        na_cols = ['Контрагент', 'Договор', 'Номер счета']
        df.reset_index(inplace=True)
        col = 'Сальдо'
        for i in range(1, len(df)):
            if all(map(lambda x: df[x][i] == None, na_cols)) and df[col][i] != None:
                for column in na_cols:
                    j = i-1
                    df[column][i] = df[column][j]
        return df



    # def edit_bank_contract(self, contract):
    #     patterns = ('дупт', 'дду', 'нупт', 'соглашение', 'переуступка', 'уст.')
    #     pattern_check = False
    #     contract = contract.lower()
    #     for pattern in patterns:
    #         if contract.find(pattern)!=-1:
    #             pattern_check =True
    #             contract = contract.replace(pattern, '').strip()
    #     contract = contract.replace('№', '')
    #     if pattern_check:
    #         for element in sorted(contract.split(' '),reverse=True):
    #             contract_pattern = self.check_pattern_in_contract(element)
    #             if contract_pattern:
    #                 return element[element.find(contract_pattern):]
    #     return contract
    #
    #
    # def check_pattern_in_contract(self, contract):
    #     patterns = ('люб', 'мп', 'зп')
    #     for pattern in patterns:
    #         if contract.find(pattern)!=-1:
    #             return pattern
    #     return False

    def split_contract(self, elem):
        if len(elem.split(' ')) > 1:
            return elem.split(' ')[1]
        else:
            return elem

    @staticmethod
    def set_index_on_df(df):
        df['bank_id'] = range(1, len(df) + 1)
        return df

    def edit_bank_agent(self, agent):
        agent = agent.replace('ё', 'е')
        if agent.isupper() and len(agent.split(' ')) == 3 and agent.startswith('ООО')==False:
            return ' '.join([elem.capitalize() for elem in agent.split(' ')])
        else:
            return agent

    @staticmethod
    def sum_amount(df):
        result = sum([float(elem) if elem != '' else 0 for elem in df['Сальдо']])
        log.info(f'Сумма по данному документу: {result}')
        return result
