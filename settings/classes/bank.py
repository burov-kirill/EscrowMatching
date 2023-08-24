import re
import warnings

import numpy as np
import openpyxl
import pandas as pd
from logs import log
from  collections import  defaultdict
from user_settings.user_exceptions import WorkbookFilterError, NotFoundColumns
from user_settings.user_interfaces import check_query_panel


class BankFile:
    COLUMNS_FOR_BANKS = {
        'СБЕР': {
            'old_columns': ['Объект строительства', 'Статус', 'Депонент', 'Номер ДДУ', 'Депонируемая сумма', 'Исходящий остаток', 'Номер счёта эскроу', 'Дата ДДУ'],
            'new_columns': ['Контрагент', 'Договор', 'Сумма по ДДУ', 'Сальдо', 'Номер счета', 'Очередь', 'Дом'],
            'status': ['Закрыт', 'Уступлен']
        },
        'МКБ': {
            'old_columns': ['ФИО плательщика', 'Номер ДДУ', 'Статус', 'Сумма по ДДУ', 'Остаток на счете эскроу', 'Номер счета эскроу', 'Дата операции'],
            'new_columns': ['Контрагент', 'Договор', 'Статус', 'Сумма по ДДУ', 'Сальдо', 'Номер счета', 'Дата операции'],
            'status': ['13', '14']
        },
        'Альфа Банк': {
            'old_columns': ['ФИО депонента', 'Данные дог.об участии сч.(ДДУ)', 'Статус счета эскроу', 'Депонируемая сумма', 'Текущий остаток на счете эскроу на дату отчета', 'Номер счета эскроу'],
            'new_columns': ['Контрагент', 'Договор', 'Статус', 'Сумма по ДДУ', 'Сальдо', 'Номер счета'],
            'status': ['закрыт']
        },
        'Совкомбанк': {
            'old_columns': ['Наименование Депонента', 'Номер ДДУ', 'Размер депонируемой суммы', 'Исходящий остаток депонируемой суммы, RUB', 'Номер счета эскроу'],
            'new_columns': ['Контрагент', 'Договор', 'Сумма по ДДУ', 'Сальдо', 'Номер счета'],
            'status': []
        },
        'ВТБ': {
            'old_columns': ['ФИО Плательщика/Наименование', 'Номер ДДУ', 'Сумма по ДДУ, руб.','Остаток на счете эскроу, руб.', 'Номер счета эскроу'],
            'new_columns': ['Контрагент', 'Договор', 'Сумма по ДДУ', 'Сальдо', 'Номер счета'],
            'status': []
        },
        'ГПБ': {
            'old_columns': ['Депонент (ФИО)', 'Номер договора основания', 'Статус счета', 'Депонируемая сумма', 'Исходящий остаток', 'Счет-эскроу', 'Номер договора счета эскроу', 'Дата операции'],
            'new_columns': ['Контрагент', 'Договор', 'Статус', 'Сумма по ДДУ', 'Сальдо', 'Номер счета', 'Эскроу', 'Дата операции'],
            'status': ['Закрыт']
        },
        'Дом РФ': {
            'old_columns': ['Депонент', '№ договора ДДУ', 'Состояние счета', 'Стоимость ДДУ, руб.', 'Остаток денежных средств на конец периода,  руб.', 'Номер счета'],
            'new_columns': ['Контрагент', 'Договор', 'Статус', 'Сумма по ДДУ', 'Сальдо', 'Номер счета'],
            'status': ['ИСПОЛНЕН']
        },
        'ПСБ': {
            'old_columns': [8, 7, 14, 6],
            'new_columns': ['Контрагент', 'Договор', 'Сальдо', 'Номер счета'],
            'rows': [0, 1, 2]
        },
        'Промсвязьбанк': {
            'old_columns': ['Наименование  депонента', 'Номер ДДУ', 'Депонируемая сумма',
                            'Исходящий остаток на эскроу счете', 'Номер счета Эскроу', "Статус счета Эскроу"],
            'new_columns': ['Контрагент', 'Договор', 'Сумма по ДДУ', 'Сальдо', 'Номер счета', "Статус"],
            'status': ['Закрыт']
        },
        'ВБРР': {
            'old_columns': ['Наименование клиента', 'Номер ДДУ', 'Дата операции', 'Сумма депонирования по ДДУ',
                            'Остаток на счете эскроу', 'Номер счета эскроу'],
            'new_columns': ['Контрагент', 'Договор','Дата операции', 'Сумма по ДДУ', 'Сальдо', 'Номер счета'],
            'status': []
        },
        'Новый ПСБ': {
            'old_columns': ['Наименование  депонента', 'Номер ДДУ', 'Дата операции ', 'Депонируемая сумма',
                            'Сумма на счете Эскроу', 'Номер счета', 'Статус счета'],
            'new_columns': ['Контрагент', 'Договор', 'Дата операции', 'Сумма по ДДУ', 'Сальдо', 'Номер счета', 'Статус'],
            'status': ['Закрыт']
        },
    }

    NEW_COLUMNS = ['Контрагент', 'Договор', 'Статус', 'Сумма по ДДУ', 'Сальдо', 'Номер счета']

    def __init__(self, user_data_dict, type_dict):
        self.contract_review = pd.DataFrame()
        self.name_bank_file = user_data_dict['bank_file']
        self.name_bank_folder = user_data_dict['bank_folder']
        if len(user_data_dict['bank_name'])>0:
            self.bank_name = user_data_dict['bank_name'][0]
        else:
            self.bank_name = ''
        self.is_one_file = user_data_dict['single']
        self.file_to_bank = user_data_dict['file_to_bank']
        self.document_sum = 0
        self.count_documents = 0
        self.type_dict = type_dict
        self.is_check = user_data_dict['check']
        self.df = self.create_and_edit_bank_dataframe()



    def create_and_edit_bank_dataframe(self):
        result = pd.DataFrame()
        if self.name_bank_file != '':
            short_file_name = self.name_bank_file[self.name_bank_file.rfind('/') + 1:]
            result = self.edit(self.bank_name, self.name_bank_file, short_file_name)
            self.count_documents+=1
        else:
            if self.bank_name != 'ВБРР':
                temp_bank_data = pd.DataFrame(columns=self.NEW_COLUMNS)
            else:
                temp_bank_data = pd.DataFrame(columns=self.COLUMNS_FOR_BANKS[self.bank_name]['old_columns'])
            for key, value in self.file_to_bank.items():
                part_bank_data = self.edit(value[1], value[0], key)
                temp_bank_data = pd.concat([temp_bank_data, part_bank_data])
                self.count_documents += 1
            result = temp_bank_data
            if self.bank_name == 'ВБРР':
                result.dropna(inplace=True)
                result = result[result['Наименование клиента'] != 'ИТОГО']
                result = self.edit_bank_data(raw_bank_data=result, bank_type=self.bank_name)
        result = self.set_index_on_df(result)
        return result

    def edit(self, bank_name, file_name, short_file_name):
        if short_file_name == 'Александров 3.xlsx':
            pass
        log.info(f'Считывание файла {short_file_name}')
        data = pd.DataFrame()
        try:
            with warnings.catch_warnings(record=True):
                warnings.simplefilter("always")
                wb = openpyxl.load_workbook(file_name)
            if bank_name == 'Дом РФ':
                raw_data = wb[wb.sheetnames[1]]
            else:
                raw_data = wb[wb.sheetnames[0]]
            data = self.edit_bank_data(raw_data, file_name=file_name, bank_type=bank_name)
        except WorkbookFilterError as exp:
            raise WorkbookFilterError(self.name_bank_file)
        else:
            wb.close()
            return data

    def edit_bank_data(self, raw_bank_data, file_name = '', bank_type = ''):
        bank_data = self.remove_na_rows(raw_bank_data, bank_type)
        if bank_type == 'ВБРР' and self.count_documents != len(self.file_to_bank) and self.bank_name != '':
            return bank_data
        try:
            bank_data = self.select_bank_columns(bank_data, bank_type, file_name)
        except NotFoundColumns as exp:
            log.exception(NotFoundColumns)
            raise NotFoundColumns
        bank_data = self.rename_columns(bank_data)
        if bank_data.empty == False and self.is_check==False:
            bank_data = self.edit_queries(bank_data)
        return bank_data

    def edit_queries(self, bank_data):
        etalon = list(set(list(map(tuple, bank_data[bank_data['Очередь'].apply(lambda x: len(str(x)) == 1)][['Очередь', 'Дом']].values.tolist()))))
        for i in range(len(bank_data)):
            for element in etalon:
                if bank_data['Очередь'][i] == element[1] and len(bank_data['Очередь'][i])>=2:
                    bank_data['Очередь'][i] = element[0]
                    bank_data['Дом'][i] = element[1]
        return bank_data

    def rename_columns(self, df):
        string_columns = ['Контрагент', 'Договор', 'Номер счета', 'Договор (полный)', 'Очередь', 'Дом']
        for column in string_columns:
            df[column] = df[column].apply(lambda x: str(x).strip())
        df["Сальдо"] = pd.to_numeric(df["Сальдо"], errors='ignore')
        df["Сумма по ДДУ"] = pd.to_numeric(df["Сумма по ДДУ"], errors='ignore')
        if 'Дата операции' in df.columns:
            df.drop(['Дата операции'], axis = 1, inplace=True)
        df = df.groupby(['Контрагент', 'Договор', 'Номер счета', 'Договор (полный)','Очередь', 'Дом'], as_index=False).agg(sum)
        self.document_sum += self.sum_amount(df)
        df['Сальдо'] = df['Сальдо'].apply(str)
        df['Контрагент'] = df['Контрагент'].apply(lambda x: self.edit_bank_agent(x))
        return df

    def remove_na_rows(self, raw_bank_data, bank):
        if self.bank_name == 'ВБРР' and self.count_documents == len(self.file_to_bank):
            return raw_bank_data
        row_list = list(raw_bank_data.values)
        bank_data = pd.DataFrame(row_list)
        col_indexes = []

        if bank == 'Совкомбанк':
            for i, col in enumerate(bank_data.columns):
                na_cols = (sum(map(lambda x: x is None, bank_data[col]))/len(bank_data[col]))*100
                if na_cols>60:
                    col_indexes.append(i)
            bank_data.drop(col_indexes, axis=1, inplace=True)

        row_indexes = []
        date_element = ''
        for i in range(len(bank_data)):
            if sum(map(lambda x: x is None or x == '', bank_data.iloc[i])) > sum(map(lambda x: x is not None and x!='', bank_data.iloc[i])) or\
                    all(map(lambda x: str(x).isdigit(), bank_data.iloc[i])):
                elements = [element for element in list(bank_data.iloc[i]) if 'За период' in str(element)]
                if len(elements) >= 1 and date_element == '':
                    date_element = elements[0]
                row_indexes.append(i)

        bank_data = bank_data.drop(index=row_indexes)
        bank_data.reset_index(inplace=True)
        date_string = re.search(r'(\d{2}\/\d{2}\/\d{4})', date_element)
        if date_string != None:
            date_string = date_string.group()
        else:
            date_string = ''

        if bank == 'ПСБ':
            bank_data.drop(index=self.COLUMNS_FOR_BANKS[bank]['rows'], inplace=True)
        else:
            bank_data.columns = bank_data.iloc[0]
            bank_data.drop(index=0, inplace=True)
            # try:
            #     bank_data.drop(index=0, inplace=True)
            # except KeyError as kerr:
            #     log.info('Данный файл пуст')
            #     bank_data.drop(index=[0], inplace=True)
            bank_data.drop(columns=bank_data.columns[0], axis=1, inplace=True)
        if bank == 'ВБРР':
            bank_data['Дата операции'] = date_string
            bank_data = bank_data[self.COLUMNS_FOR_BANKS[bank]['old_columns']]
        return bank_data

    def select_bank_columns(self, df, bank, filename):
        df.columns = df.columns.map(lambda x: x.replace('\n', ' ') if type(x) == str else x)
        if bank in ('МКБ', 'ГПБ') and all(map(lambda x: 'Статус' not in x,  df.columns)):
            if bank == 'МКБ':
                df['Статус'] = ''
            else:
                df['Статус счета'] = ''
        df = df[self.COLUMNS_FOR_BANKS[bank]['old_columns']]

        if bank == 'СБЕР':
            df = df.query(f'Статус not in {self.COLUMNS_FOR_BANKS[bank]["status"]}')
            df = self.set_query_and_house(df, bank, filename)

            self.contract_review = self.create_check_correct_values(df)
            df = df[['Депонент', 'Номер ДДУ', 'Депонируемая сумма', 'Исходящий остаток', 'Номер счёта эскроу', 'Очередь', 'Дом']]
            df.columns = self.COLUMNS_FOR_BANKS[bank]['new_columns']
        elif bank in ('МКБ', 'ГПБ'):
            df.columns = self.COLUMNS_FOR_BANKS[bank]['new_columns']
            df['Договор'] = df['Договор'].apply(self.edit_bank_contract)
        else:
            df.columns = self.COLUMNS_FOR_BANKS[bank]['new_columns']
            df['Договор'] = df['Договор'].apply(self.edit_bank_contract)
            if 'Сумма по ДДУ' not in df.columns:
                df['Сумма по ДДУ'] = 0

            if bank in ('Альфа Банк', 'Новый ПСБ'):
                df = df.query(f'Статус not in {self.COLUMNS_FOR_BANKS[bank]["status"]}')
            if bank == 'Совкомбанк' or 'Совкомбанк' in list(map(lambda x: x[1], self.file_to_bank.values())):
                df['Контрагент'] = df['Контрагент'].apply(self.edit_account_agent)
        if bank == 'МКБ':
            df['Сальдо'] = df['Сальдо'].apply(lambda x: float(str(x).replace(u'\xa0', u'').replace(',', '.')))
            df['Сумма по ДДУ'] = df['Сумма по ДДУ'].apply(lambda x: float(str(x).replace(u'\xa0', u'').replace(',', '.')))
        elif bank == 'Дом РФ':
            df = df[df['Сальдо'].map(lambda x: type(x) != str)]
            df = df.query(f'Статус not in {self.COLUMNS_FOR_BANKS[bank]["status"]}')
            df = self.edit_DOM_contract(df, filename)

        df['Договор (полный)'] = df['Договор'].apply(lambda x: str(x).upper())
        if bank!='СБЕР':
            df['Очередь'] = df['Договор'].apply(self.find_queries)
            df['Дом'] = df['Договор'].apply(self.find_queries, args=[False])
            df['Индекс'] = df['Очередь'] + df['Дом']
            df = self.set_query_and_house(df, bank, filename)
            # if self.is_check:
            #     df['Индекс'] = df['Очередь'] + df['Дом']
            #     query_house_dict = self.create_query_house_dict(df, bank)
            #     correct_values_dict = check_query_panel(query_house_dict, bank, filename)
            #     df['Очередь'] = df['Индекс'].apply(self.set_correct_query, args=[correct_values_dict])
            #     df['Дом'] = df['Индекс'].apply(self.set_correct_house, args=[correct_values_dict])
        if bank in ('МКБ', 'ГПБ', 'ВБРР'):
            df = self.drop_double_rows(df, bank)
            if bank != 'ВБРР':
                df = df.query(f'Статус not in {self.COLUMNS_FOR_BANKS[bank]["status"]}')

        df['Договор'] = df['Договор'].apply(lambda x: str(x).upper())
        df['Договор'] = df['Договор'].apply(self.clear_contract)
        df['Контрагент'] = df['Контрагент'].apply(self.clear_agent)
        return df
    @staticmethod
    def clear_agent(agent):
        if isinstance(agent, str):
            return agent.replace('ё', 'е').replace('Ё', 'Е')
        else:
            return agent
    @staticmethod
    def clear_contract(contract):
        if isinstance(contract, str):
            return contract.replace(' ', '')
        else:
            return contract
    def edit_account_agent(self, agent):
        agent_list = str(agent).split(' ')
        if len(agent_list) == 3:
            return f'{agent_list[0].upper()} {agent_list[1][0]}.{agent_list[2][0]}.'
        else:
            return agent
    def set_query_and_house(self, df, bank, filename):
        column = 'Объект строительства' if bank == 'СБЕР' else 'Индекс'
        if self.is_check:
            query_house_dict = self.create_query_house_dict(df, bank)
            correct_values_dict = check_query_panel(query_house_dict, bank, filename)
            df['Очередь'] = df[column].apply(self.set_correct_query, args=[correct_values_dict])
            df['Дом'] = df[column].apply(self.set_correct_house, args=[correct_values_dict])
        else:
            if bank == 'СБЕР':
                df['Очередь'] = df[column].apply(self.get_query)
                df['Дом'] = df[column].apply(self.get_query, args=[False])

        return df

    def create_check_correct_values(self, df):
        df['Очередь (Договор)'] = df['Номер ДДУ'].apply(self.find_queries)
        df['Дом (Договор)'] = df['Номер ДДУ'].apply(self.find_queries, args=[False])
        df = df[['Объект строительства', 'Номер ДДУ', 'Дата ДДУ', 'Исходящий остаток', 'Очередь', 'Дом', 'Очередь (Договор)', 'Дом (Договор)']]
        df.columns = ['Объект строительства', 'Договор', 'Дата договора', 'Остаток', 'Очередь (ОС)', 'Дом (ОС)', 'Очередь (Договор)', 'Дом (Договор)']
        df['Проверка Очереди'] = np.where(df['Очередь (ОС)'] == df['Очередь (Договор)'], 'ИСТИНА', 'ЛОЖЬ')
        df['Проверка Дома'] = np.where(df['Дом (ОС)'] == df['Дом (Договор)'], 'ИСТИНА', 'ЛОЖЬ')
        return df


    def drop_double_rows(self, df, bank):
        values_for_bank = {'МКБ': ['Контрагент','Договор','Номер счета', 'Дата операции', 'Сальдо'],
                           'ВБРР': ['Контрагент','Договор','Номер счета', 'Дата операции', 'Сальдо'],
                        'ГПБ': ['Контрагент', 'Договор', 'Эскроу', 'Дата операции']}
        df['Дата операции'] = pd.to_datetime(df['Дата операции'], errors='coerce', dayfirst=True)
        df['Номер счета'] = df['Номер счета'].apply(str)
        df.sort_values(by=values_for_bank[bank], ascending=False,
                       inplace=True)
        df.reset_index(inplace=True)
        df['Help_Col'] = 0
        res = []
        if bank in ('МКБ', 'ВБРР'):
            for i in range(len(df)):
                if (df['Договор'][i], df['Номер счета'][i]) not in res:
                    df['Help_Col'][i] = 1
                    res.append((df['Договор'][i], df['Номер счета'][i]))
        elif bank == 'ГПБ':
            for i in range(len(df)):
                if (df['Договор'][i], df['Эскроу'][i]) not in res and len(df.iloc[i:].loc[(df['Номер счета']==df['Номер счета'][i]) & (df['Договор']!=df['Договор'][i])])==0:
                    df['Help_Col'][i] = 1
                    res.append((df['Договор'][i], df['Эскроу'][i]))
        df = df[df['Help_Col'] == 1]
        return df




    def edit_DOM_contract(self, df, filename):
        wb = openpyxl.load_workbook(filename)
        raw_data = wb[wb.sheetnames[2]]
        new_data = self.remove_na_rows(raw_data, 'Дом РФ')
        new_data['Новый договор'] = new_data['Назначение платежа'].apply(self.edit_bank_contract)
        new_data.reset_index(inplace=True)
        df.reset_index(inplace=True)
        result = defaultdict(list)
        for i in range(len(new_data)):
            key_str = f"{new_data['Депонент'][i]}_{new_data['Счет'][i]}"
            result[key_str].append(new_data['Новый договор'][i])
            # result.setdefault(key_str, []).append(new_data['Новый договор'][i])
        df['Договор'] = df.apply(lambda x: result.get(f"{x['Контрагент']}_{x['Номер счета']}",
                                                                ['бн'])[-1] if x['Договор'] == '' else x['Договор'], axis=1)
        # for i in range(len(df)):
        #     if df['Договор'][i] == '':
        #         df['Договор'][i] = result.get(f"{df['Контрагент'][i]}_{df['Номер счета'][i]}", ['бн'])[-1]
        return df

    @staticmethod
    def set_correct_query(stirng, correct_dict):
        return correct_dict[stirng][0]

    @staticmethod
    def set_correct_house(stirng, correct_dict):
        return correct_dict[stirng][1]

    def create_query_house_dict(self, df, bank_name):
        if bank_name=='СБЕР':
            query_column = 'Объект строительства'
            df['Очередь'] = df[query_column].apply(self.get_query)
            df['Дом'] = df[query_column].apply(self.get_query, args=[False])
        else:
            query_column = 'Индекс'
        columns_list = [query_column, 'Очередь', 'Дом']
        df = df.groupby(columns_list, as_index=False).count()
        result = {df[query_column][i]: (df['Очередь'][i], df['Дом'][i]) for i in range(len(df))}
        return result
        #     # result = dict()
        #     df['Очередь'] = df['Объект строительства'].apply(self.get_query)
        #     df['Дом'] = df['Объект строительства'].apply(self.get_query, args=[False])
        #     df = df.groupby(['Объект строительства', 'Очередь', 'Дом'], as_index=False).count()
        #     result = {df['Объект строительства'][i]: (df['Очередь'][i],df['Дом'][i]) for i in range(len(df))}
        #     # for i in range(len(df)):
        #     #     result[df['Объект строительства'][i]] = (df['Очередь'][i],df['Дом'][i])
        #     return result
        # else:
        #     # result = dict()
        #     df = df.groupby(['Индекс', 'Очередь', 'Дом'], as_index=False).count()
        #     result = {df['Индекс'][i]: (df['Очередь'][i], df['Дом'][i]) for i in range(len(df))}
        #     # for i in range(len(df)):
        #     #     result[df['Индекс'][i]] = (df['Очередь'][i], df['Дом'][i])
        #     return result

    def find_queries(self, string, option=True):
        pattern = r'(?<!\d-)(\d+[.,]?\d{0,2})'
        old_pattern = r'\d+[,.-]?\d{0,2}'
        numbers = re.findall(pattern, string)
        if len(numbers) >= 2:
            if option:
                return numbers[0]
            else:
                return numbers[1]
        else:
            return 'бн'

    def get_query(self, string, option = True):
        pattern = r'\d+[,.-]?\d{0,2}'
        lst = re.findall(pattern, string)[::-1]
        if len(lst)>0:
            for element in lst:
                if element in list(self.type_dict.keys()):
                    if option:
                        return self.type_dict[element]
                    else:
                        return element
            return lst[-1]
        else:
            return string




    # def get_query(self, string, option = True):
    #     pattern = {1: '№ (\d+)',2: r'(\d+) дом',3:r'№(\d+\w\d*)', 4:r'№(\d*)', 5: r'корпус[а]? (\d\.?\d?)', 6: r'дом[а]? (\d+)', 7:r'корп\.? (\d+\.?\d?)'}
    #     candidat = ''
    #     for key, value in pattern.items():
    #         lst = re.findall(value, string)
    #         if len(lst)>0:
    #             if key == 3:
    #                 candidat =  lst[0].replace('к','.')
    #             else:
    #                 candidat =  lst[0]
    #             if candidat in list(self.type_dict.keys()):
    #                 if option:
    #                     return self.type_dict[candidat]
    #                 else:
    #                     return candidat
    #             else:
    #                 return candidat
    #     return string


    def fill_na_bank_data(self, df):
        na_cols = ['Контрагент', 'Договор', 'Номер счета']
        df.reset_index(inplace=True)
        col = 'Сальдо'
        for i in range(1, len(df)):
            if all(map(lambda x: df[x][i] == None, na_cols)) and df[col][i] != None:
                for column in na_cols:
                    j = i-1
                    df[column][i] = df[column][j]
        return df

    def edit_bank_contract(self, contract):
        pattern = r'[A-Za-zА-Яа-я]+[-/]\d+[-/]\d*[.,-]?\d*[-/]?\d*[-/]?[А-Яа-я0-9\-]*'
        match = re.findall(pattern, contract)
        if len(match) >= 1:
            for element in match:
                numbers = re.findall(r'\d+[.-]?\d{0,2}', element)
                if len(numbers[0])==1 or numbers[0].find('.')!=-1:
                    return element
            return match[-1]
        return contract


    def split_contract(self, elem):
        if len(elem.split(' ')) > 1:
            return elem.split(' ')[1]
        else:
            return elem

    @staticmethod
    def set_index_on_df(df):
        df['bank_id'] = range(1, len(df) + 1)
        return df

    def edit_bank_agent(self, agent):
        agent = agent.replace('ё', 'е')
        if agent.isupper() and len(agent.split(' ')) == 3 and agent.startswith('ООО')==False:
            return ' '.join([elem.capitalize() for elem in agent.split(' ')])
        else:
            return agent

    @staticmethod
    def sum_amount(df):
        result = sum([float(elem) if elem != '' else 0 for elem in df['Сальдо']])
        log.info(f'Сумма по данному документу: {result}')
        return result
