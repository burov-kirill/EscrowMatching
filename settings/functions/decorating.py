import openpyxl
import pandas as pd
from openpyxl.formatting import Rule
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.colors import BLUE

from logs import log

description = {1: 'Расхождения по всем четырем столбцам (ФИО, Договор, Номер счета Эскроу, Сумма)',
               2: 'Расхождения по столбцам «Контрагент» и «Договор»',
               3: 'Расхождения по столбцам «Договор» и «Номер счета»',
               4: 'Расхождения по столбцам «Договор» и «Сальдо»',
               5: 'Расхождения по столбцам «Сальдо» и «Номер счета»',
               6: 'Расхождения по столбцам «Номер счета» и «Контрагент»',
               7: 'Расхождения по столбцам «Сальдо» и «Контрагент»',
               8: 'Расхождения по столбцу «Номер счета»',
               9: 'Расхождения по столбцу «Договор»',
               10: 'Расхождения по столбцу «Сальдо»',
               11: 'Расхождения по столбцу «Контрагент»',
               12: 'Данные сходятся по всем столбцам',
               }

TABLE_NUMBER = 1
def as_text(value):
    if value is None:
        return ""
    return str(value)

def auto_width_columns(ws, opt = True):
    columns = {2: 22,
               3: 27,
               4: 20,
               5: 18,
               6: 4,
               7: 12,
               8: 4,
               9: 27,
               10: 25,
               11: 20,
               12: 18}
    if opt:
        for column_cells in ws.columns:
            length = max(len(as_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5
    else:
        for col, width in columns.items():
            ws.column_dimensions[get_column_letter(col)].width = width
def decoration(result_dict, review, review_for_MSFO, one_more_review, contract_review, path):
    log.info(f'Оформление результирующего листа')
    control_bank_sum = 0
    control_account_sum = 0
    wb = openpyxl.Workbook()
    wb.guess_types = True
    review_sheet = wb[wb.sheetnames[0]]
    review_sheet.title = 'Ревью'
    decoration_table(review_sheet, review, 4)
    decoration_table(review_sheet, review_for_MSFO, 10)
    decoration_table(review_sheet, one_more_review, 16)
    auto_width_columns(review_sheet)

    if contract_review.empty == False:
        check_sheet = wb.create_sheet('Проверка')
        decoration_table(check_sheet, contract_review, 2)

    for key, value in sorted(result_dict.items(), reverse=True):
        ws = wb.create_sheet(str(key))
        temp_bank_sum, temp_account_sum = excel_list_decoration(value, ws)
        if key!='Общий':
            control_account_sum+=temp_account_sum
            control_bank_sum+=temp_bank_sum

    name = 'Сверка.xlsx'
    path = f'{path}/{name}'
    wb.save(path)
    wb.close()


    return (control_bank_sum, control_account_sum)

def decoration_table(ws, table, col):
    rows = dataframe_to_rows(table, index=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, col):
            if c_idx == 14:
                ws.cell(row=r_idx, column=c_idx, value=value).number_format = '0.00%'
            else:
                ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'

    init_col = get_column_letter(col)
    end_col = get_column_letter(len(table.columns)+col-1)
    table_length = len(table) + 2
    global TABLE_NUMBER
    tab = Table(displayName=f"Table{TABLE_NUMBER}",
                ref=f"{init_col}2:{end_col}{table_length}")
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=True,
                           showLastColumn=True, showRowStripes=False, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    TABLE_NUMBER+=1
    #
def excel_list_decoration(d, ws):

    control_bank_sum = 0
    control_account_sum = 0
    merge_cell = set_value_cell(ws, 'B1:F1', 'БАНК') # исправить
    header_decoration(merge_cell, 'head')

    merge_cell = set_value_cell(ws, 'J1:N1', '1C') # исправить
    header_decoration(merge_cell, 'head')

    ws['H1'] = 'Разница' # исправить
    ws['H2'] = '=N2-F2' # исправить
    ws['H2'].border = Border(left=Side(style='thick'), right=Side(style='thick'),
                             top=Side(style='thick'), bottom=Side(style='thick'))
    header_decoration(ws['H1'], 'head') # исправить
    header_decoration(ws['H2'], 'sub') # исправить

    set_value_cell(ws, ('B2:E2', 'J2:M2'), 'ИТОГО по очереди и дому') # исправить
    set_value_cell(ws, ('B3', 'J3'), 'Номер счета/дома') # исправить
    set_value_cell(ws, ('C3', 'K3'), 'Номенклатурная группа')  # исправить
    set_value_cell(ws, ('D3', 'L3'), 'ФИО депонента') # исправить
    set_value_cell(ws, ('E3', 'M3'), 'Данные дог.об участии сч.(ДДУ)') # исправить
    set_value_cell(ws, ('F3', 'N3'), 'остаток') # исправить

    part_bank_sum, part_account_sum = [], []

    # Проходимся по словарю и заносим фреймы в лист
    for key, value in sorted(d.items()):
        last_row = len(ws['A']) + 5  # Находим последнюю ячейку
        merge_cell = set_value_cell(ws, f'B{last_row}:N{last_row + 1}',
                                    description[key])  # Объединяем и заполянем строку с названием блока # исправить!!!
        header_decoration(merge_cell, 'block')  # Оформляем строку с названием блока
        set_value_cell(ws, (f'B{last_row + 2}:E{last_row + 2}', f'J{last_row + 2}:M{last_row + 2}'),
                       f'ИТОГО_часть{key}')  # Заполняем строку с данными под блоком # исправить!!!
        part_bank_sum.append(f'F{last_row + 2}')  # Добавляем список имя ячейки где будет храниться сумма блока по банку # исправить!!
        part_account_sum.append(f'N{last_row + 2}') # исправить!!!
        ws[f'H{last_row+2}'] = f'=N{last_row + 2}-F{last_row + 2}' # исправить
        ws[f'H{last_row + 2}'].number_format = '#,##0.00' # исправить
        ws[f'H{last_row + 2}'].alignment = Alignment(horizontal="center", vertical="center") # исправить
        ws[f'H{last_row + 2}'].font = Font(bold=True, color="000000", name='Calibri', size=12) # исправить
        # Добавляем список имя ячейки где будет храниться сумма блока по 1С
        bank_df, account_df = split_df(key, value)  # Разбиваем фрейм на банк и 1С
        past_data_frame_to_excel_list(ws, bank_df, last_row + 3, True)  # Заносим фрейм в лист
        past_data_frame_to_excel_list(ws, account_df, last_row + 3, False)
        auto_width_columns(ws, False)

        control_bank_sum += bank_df['Сальдо'].sum()
        control_account_sum += account_df['Сальдо'].sum()

    draw_border_for_bottom_line(ws, 'F2', part_bank_sum)  # Рисуем границу для верхних итогов # исправить
    draw_border_for_bottom_line(ws, 'N2', part_account_sum) # исправить




    return (control_bank_sum, control_account_sum)


# Заполняем и оформляем верхние итоги
def draw_border_for_bottom_line(ws, cell, cell_list):
    ws[cell] = f'={"+".join(cell_list)}'
    ws[cell].number_format = '#,##0.00'
    ws[cell].border = Border(left=Side(style='thick'), right=Side(style='thick'),
                             top=Side(style='thick'), bottom=Side(style='thick'))
    header_decoration(ws[cell], 'sub')

    # Рисуем границы объединенных ячеек
    for merged_cells in ws.merged_cells.ranges:
        for col in range(merged_cells.min_col, merged_cells.max_col + 1):
            for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                ws.cell(row, col).border = Border(left=Side(style='thick'), right=Side(style='thick'),
                                                  top=Side(style='thick'), bottom=Side(style='thick'))


# Оформление заголовков
def header_decoration(cell, option):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if option == 'head':
        cell.fill = PatternFill("solid", fgColor="002060")  # Изменить цвет
        cell.font = Font(bold=True, color="FFFFFF", name='Calibri', size=14)
    elif option == 'block':
        cell.fill = PatternFill("solid", fgColor="FFC000")  # Изменить цвет
        cell.font = Font(bold=True, color="000000", name='Calibri', size=14)
    else:
        cell.fill = PatternFill("solid", fgColor="FFFFFF")  # Изменить цвет
        cell.font = Font(bold=True, color="000000", name='Calibri', size=12)


# Заполнение ячеек. Если передана строка, то возвращается объединенная строка,
# если же передан кортеж, то просто оформляется ячейка
def set_value_cell(ws, cell_range, value):
    if isinstance(cell_range, str):
        ws.merge_cells(cell_range)
        merge_cell = ws[cell_range.split(':')[0]]
        merge_cell.value = value
        merge_cell.number_format = '#,##0.00'
        header_decoration(merge_cell, 'sub')
        return merge_cell
    else:
        for item in cell_range:
            ws.merge_cells(item)
            merge_cell = ws[item.split(':')[0]]
            merge_cell.value = value
            merge_cell.number_format = '#,##0.00'
            header_decoration(merge_cell, 'sub')


# Перенос фрейма данных в лист
def past_data_frame_to_excel_list(ws, df: pd.DataFrame, cell_row: int, option: bool):
    rows = dataframe_to_rows(df, index=False)
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="duplicateValues", text="highlight", dxf=dxf)
    if option:
        for r_idx, row in enumerate(rows, cell_row):
            for c_idx, value in enumerate(row, 2):
                ws.cell(row=r_idx, column=c_idx, value=value)
                ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'
        set_value_cell(ws, f'F{cell_row - 1}',
                       f'=SUM(F{cell_row + 1}:F{cell_row + len(df) + 1})') # исправить

        ws.conditional_formatting.add(f'D{cell_row + 1}:D{cell_row + len(df) + 1}', rule)
        # Заполняем промежуточные итоги
    else:
        for r_idx, row in enumerate(rows, cell_row):
            for c_idx, value in enumerate(row, 10):
                if str(value).startswith('!BLUE'):
                    value = value.split('_')[1]
                    ws.cell(row=r_idx, column=c_idx, value=value).fill = PatternFill(fgColor='00BFFF', fill_type = "solid")
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'
                # ws.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

        set_value_cell(ws, f'N{cell_row - 1}', f'=SUM(N{cell_row + 1}:N{cell_row + len(df) + 1})') # исправить
        ws.conditional_formatting.add(f'L{cell_row + 1}:L{cell_row + len(df) + 1}', rule) # исправить

def split_df(key, tuple_df):
    bank = tuple_df[0]
    account = tuple_df[1]
    bank['Тип'] = bank['Очередь'] + '_' + bank['Дом']
    bank = bank[['Номер счета', 'Тип', 'Контрагент', 'Договор (полный)', 'Сальдо', ]]
    account = account[['Номер счета', 'Тип', 'Контрагент', 'Договор', 'Сальдо']]
    bank["Сальдо"] = pd.to_numeric(bank["Сальдо"], errors='ignore')
    account["Сальдо"] = pd.to_numeric(account["Сальдо"], errors='ignore')
    if key == 2:
        bank.sort_values(by=['Контрагент'], inplace=True)
        account.sort_values(by=['Контрагент'], inplace=True)

    else:
        bank.sort_values(by=['Договор (полный)'], inplace=True)
        account.sort_values(by=['Договор'], inplace=True)
    bank.sort_values(by=['Контрагент', 'Договор (полный)'], inplace=True)
    account.sort_values(by=['Контрагент', 'Договор'], inplace=True)
    bank.dropna(axis=0, inplace=True)
    # bank = bank.loc[(bank['Сальдо']!=0)]
    return (bank, account)

