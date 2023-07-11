import openpyxl
import pandas as pd
from openpyxl.formatting import Rule
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwings as xw
from openpyxl.worksheet.table import Table, TableStyleInfo

from logs import log

description = {1: 'Расхождения по всем трём столбцам (ФИО, Договор, Сумма)',
               2: 'Расхождения по столбцу «Договор»',
               3: 'Расхождения по столбцу «Сумма»',
               4: 'Расхождения по столбцу «ФИО» ',
               5: 'Данные сходятся по всем трём столбцам'}

def decoration(result_dict, review, review_for_MSFO, contract_review, path):
    log.info(f'Оформление результирующего листа')
    control_bank_sum = 0
    control_account_sum = 0
    wb = openpyxl.Workbook()
    wb.guess_types = True
    review_sheet = wb[wb.sheetnames[0]]
    review_sheet.title = 'Ревью'
    decoration_table(review_sheet, review, 4, 1)
    decoration_table(review_sheet, review_for_MSFO, 10, 2)
    if contract_review.empty == False:
        check_sheet = wb.create_sheet('Проверка')
        decoration_table(check_sheet, contract_review, 2, 3)
    # rows = dataframe_to_rows(review, index=False)
    #
    #
    # for r_idx, row in enumerate(rows, 2):
    #     for c_idx, value in enumerate(row, 4):
    #         review_sheet.cell(row=r_idx, column=c_idx, value=value)
    #
    # tab = Table(displayName="Table1", ref=f"D2:G{len(review)+2}")  # Name Manager
    # style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=True,
    #                        showLastColumn=True, showRowStripes=False, showColumnStripes=True)
    # tab.tableStyleInfo = style
    # review_sheet.add_table(tab)
    #
    # rows = dataframe_to_rows(review_for_MSFO, index=False)
    #
    # for r_idx, row in enumerate(rows, 2):
    #     for c_idx, value in enumerate(row, 10):
    #         review_sheet.cell(row=r_idx, column=c_idx, value=value)
    #
    # tab = Table(displayName="Table1", ref=f"J2:G{len(review_for_MSFO) + 2}")  # Name Manager
    # style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=True,
    #                        showLastColumn=True, showRowStripes=False, showColumnStripes=True)
    # tab.tableStyleInfo = style
    # review_sheet.add_table(tab)

    for key, value in sorted(result_dict.items(), reverse=True):
        ws = wb.create_sheet(str(key))
        temp_bank_sum, temp_account_sum = excel_list_decoration(value, ws)
        if key!='Общий':
            control_account_sum+=temp_account_sum
            control_bank_sum+=temp_bank_sum

    name = 'Сверка.xlsx'
    path = f'{path}/{name}'
    wb.save(path)
    wb.close()


    return (control_bank_sum, control_account_sum)

def decoration_table(ws, table, col, number):
    rows = dataframe_to_rows(table, index=False)

    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, col):
            # ws.cell(row=r_idx, column=c_idx, value=value)
            ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'

    if number==1:
        tab = Table(displayName=f"Table{number}", ref=f"D2:H{len(table) + 2}")  # Name Manager
    elif number==2:
        tab = Table(displayName=f"Table{number}", ref=f"J2:L{len(table) + 2}")  # Name Manager
    else:
        tab = Table(displayName=f"Table{number}", ref=f"B2:J{len(table) + 2}")  # Name Manager

    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=True,
                           showLastColumn=True, showRowStripes=False, showColumnStripes=True)
    tab.tableStyleInfo = style
    ws.add_table(tab)


def excel_list_decoration(d, ws):

    control_bank_sum = 0
    control_account_sum = 0
    merge_cell = set_value_cell(ws, 'B1:E1', 'БАНК')
    header_decoration(merge_cell, 'head')

    merge_cell = set_value_cell(ws, 'I1:L1', '1C')
    header_decoration(merge_cell, 'head')

    ws['G1'] = 'Разница'
    ws['G2'] = '=L2-E2'
    ws['G2'].border = Border(left=Side(style='thick'), right=Side(style='thick'),
                             top=Side(style='thick'), bottom=Side(style='thick'))
    header_decoration(ws['G1'], 'head')
    header_decoration(ws['G2'], 'sub')

    set_value_cell(ws, ('B2:D2', 'I2:K2'), 'ИТОГО по очереди и дому')
    set_value_cell(ws, ('B3', 'I3'), 'Номер счета/дома')
    set_value_cell(ws, ('C3', 'J3'), 'ФИО депонента')
    set_value_cell(ws, ('D3', 'K3'), 'Данные дог.об участии сч.(ДДУ)')
    set_value_cell(ws, ('E3', 'L3'), 'остаток')

    part_bank_sum, part_account_sum = [], []

    # Проходимся по словарю и заносим фреймы в лист
    for key, value in sorted(d.items()):
        last_row = len(ws['A']) + 5  # Находим последнюю ячейку
        merge_cell = set_value_cell(ws, f'B{last_row}:L{last_row + 1}',
                                    description[key])  # Объединяем и заполянем строку с названием блока
        header_decoration(merge_cell, 'block')  # Оформляем строку с названием блока
        set_value_cell(ws, (f'B{last_row + 2}:D{last_row + 2}', f'I{last_row + 2}:K{last_row + 2}'),
                       f'ИТОГО_часть{key}')  # Заполняем строку с данными под блоком
        part_bank_sum.append(f'E{last_row + 2}')  # Добавляем список имя ячейки где будет храниться сумма блока по банку
        part_account_sum.append(f'L{last_row + 2}')
        ws[f'G{last_row+2}'] = f'=L{last_row + 2}-E{last_row + 2}'
        ws[f'G{last_row + 2}'].number_format = '#,##0.00'
        ws[f'G{last_row + 2}'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'G{last_row + 2}'].font = Font(bold=True, color="000000", name='Calibri', size=12)
        # Добавляем список имя ячейки где будет храниться сумма блока по 1С
        bank_df, account_df = split_df(key, value)  # Разбиваем фрейм на банк и 1С
        past_data_frame_to_excel_list(ws, bank_df, last_row + 3, True)  # Заносим фрейм в лист
        past_data_frame_to_excel_list(ws, account_df, last_row + 3, False)
        control_bank_sum += bank_df['Сальдо'].sum()
        control_account_sum += account_df['Сальдо'].sum()

    draw_border_for_bottom_line(ws, 'E2', part_bank_sum)  # Рисуем границу для верхних итогов
    draw_border_for_bottom_line(ws, 'L2', part_account_sum)




    return (control_bank_sum, control_account_sum)


# Заполняем и оформляем верхние итоги
def draw_border_for_bottom_line(ws, cell, cell_list):
    ws[cell] = f'={"+".join(cell_list)}'
    ws[cell].number_format = '#,##0.00'
    ws[cell].border = Border(left=Side(style='thick'), right=Side(style='thick'),
                             top=Side(style='thick'), bottom=Side(style='thick'))
    header_decoration(ws[cell], 'sub')

    # Рисуем границы объединенных ячеек
    for merged_cells in ws.merged_cells.ranges:
        for col in range(merged_cells.min_col, merged_cells.max_col + 1):
            for row in range(merged_cells.min_row, merged_cells.max_row + 1):
                ws.cell(row, col).border = Border(left=Side(style='thick'), right=Side(style='thick'),
                                                  top=Side(style='thick'), bottom=Side(style='thick'))


# Оформление заголовков
def header_decoration(cell, option):
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if option == 'head':
        cell.fill = PatternFill("solid", fgColor="002060")  # Изменить цвет
        cell.font = Font(bold=True, color="FFFFFF", name='Calibri', size=14)
    elif option == 'block':
        cell.fill = PatternFill("solid", fgColor="FFC000")  # Изменить цвет
        cell.font = Font(bold=True, color="000000", name='Calibri', size=14)
    else:
        cell.fill = PatternFill("solid", fgColor="FFFFFF")  # Изменить цвет
        cell.font = Font(bold=True, color="000000", name='Calibri', size=12)


# Заполнение ячеек. Если передана строка, то возвращается объединенная строка,
# если же передан кортеж, то просто оформляется ячейка
def set_value_cell(ws, cell_range, value):
    if isinstance(cell_range, str):
        ws.merge_cells(cell_range)
        merge_cell = ws[cell_range.split(':')[0]]
        merge_cell.value = value
        merge_cell.number_format = '#,##0.00'
        header_decoration(merge_cell, 'sub')
        return merge_cell
    else:
        for item in cell_range:
            ws.merge_cells(item)
            merge_cell = ws[item.split(':')[0]]
            merge_cell.value = value
            merge_cell.number_format = '#,##0.00'
            header_decoration(merge_cell, 'sub')


# Перенос фрейма данных в лист
def past_data_frame_to_excel_list(ws, df: pd.DataFrame, cell_row: int, option: bool):
    rows = dataframe_to_rows(df, index=False)
    red_text = Font(color="9C0006")
    red_fill = PatternFill(bgColor="FFC7CE")
    dxf = DifferentialStyle(font=red_text, fill=red_fill)
    rule = Rule(type="duplicateValues", text="highlight", dxf=dxf)
    if option:
        for r_idx, row in enumerate(rows, cell_row):
            for c_idx, value in enumerate(row, 2):
                ws.cell(row=r_idx, column=c_idx, value=value)
                ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'
        set_value_cell(ws, f'E{cell_row - 1}',
                       f'=SUM(E{cell_row + 1}:E{cell_row + len(df) + 1})')

        ws.conditional_formatting.add(f'C{cell_row + 1}:C{cell_row + len(df) + 1}', rule)
        # Заполняем промежуточные итоги
    else:
        for r_idx, row in enumerate(rows, cell_row):
            for c_idx, value in enumerate(row, 9):
                ws.cell(row=r_idx, column=c_idx, value=value).number_format = '#,##0.00'
                # ws.cell(row=r_idx, column=c_idx).number_format = '#,##0.00'

        set_value_cell(ws, f'L{cell_row - 1}', f'=SUM(L{cell_row + 1}:L{cell_row + len(df) + 1})')
        ws.conditional_formatting.add(f'J{cell_row + 1}:J{cell_row + len(df) + 1}', rule)

def split_df(key, tuple_df):
    bank = tuple_df[0]
    account = tuple_df[1]
    bank = bank[['Номер счета', 'Контрагент', 'Договор (полный)', 'Сальдо', ]]
    account = account[['Тип','Контрагент', 'Договор', 'Сальдо']]
    bank["Сальдо"] = pd.to_numeric(bank["Сальдо"], errors='ignore')
    account["Сальдо"] = pd.to_numeric(account["Сальдо"], errors='ignore')
    if key == 2:
        bank.sort_values(by=['Контрагент'], inplace=True)
        account.sort_values(by=['Контрагент'], inplace=True)

    else:
        bank.sort_values(by=['Договор (полный)'], inplace=True)
        account.sort_values(by=['Договор'], inplace=True)
    bank.sort_values(by=['Договор (полный)', 'Контрагент'], inplace=True)
    account.sort_values(by=['Тип','Договор', 'Контрагент'], inplace=True)
    bank.dropna(axis=0, inplace=True)
    bank = bank.loc[(bank['Сальдо']!=0)]
    return (bank, account)

